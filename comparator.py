from __future__ import annotations

from dataclasses import dataclass, field
from difflib import SequenceMatcher
from io import BytesIO
from typing import BinaryIO
import re
import unicodedata

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


MONEY_TOLERANCE = 0.05

FILL_TITLE = PatternFill("solid", fgColor="1F4E78")
FILL_SECTION = PatternFill("solid", fgColor="D9EAF7")
FILL_HEADER = PatternFill("solid", fgColor="BDD7EE")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_DIVERGENT = PatternFill("solid", fgColor="F4CCCC")
FILL_MISSING = PatternFill("solid", fgColor="FCE5CD")
FILL_FOUND = PatternFill("solid", fgColor="D9E2F3")
FILL_NOT_FOUND_DRE = PatternFill("solid", fgColor="EAD1DC")
FILL_NEUTRAL = PatternFill("solid", fgColor="EDEDED")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
FONT_BOLD = Font(bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

STOPWORDS = {
    "A",
    "AS",
    "C",
    "COM",
    "DA",
    "DAS",
    "DE",
    "DO",
    "DOS",
    "E",
    "EM",
    "NO",
    "NOS",
    "NA",
    "NAS",
    "O",
    "OS",
    "PARA",
    "POR",
    "S",
}

EXCLUDED_RESULT_PREFIXES = ("311", "312", "324", "325", "326")

EXCLUDED_KEYWORDS = {
    "13 SALARIO",
    "ADIANTAMENTO",
    "ADICIONAL NOTURNO",
    "ALIMENTACAO",
    "AMORTIZ",
    "AVISO PREVIO",
    "BENEFICIO",
    "BOLSA",
    "CAJU",
    "CONSIGNADO",
    "CSLL",
    "DEPRECI",
    "DESONERACAO",
    "EMPRESTIMO",
    "ENCARGOS",
    "ESTAGIARIO",
    "FERIAS",
    "FGTS",
    "FINANCEIR",
    "FOLHA",
    "GRATIFIC",
    "HORA EXTRA",
    "IMPOSTO",
    "INSS",
    "IOF",
    "IRPJ",
    "ISS",
    "JUROS",
    "PIS",
    "PLANO DE SAUDE",
    "PREMIO",
    "PREMIO E PRODUTIVIDADE",
    "PROVISAO",
    "RECUPERACAO",
    "RECRUTAMENTO",
    "RENDIMENTO",
    "RESCISAO",
    "SALARIO",
    "SEGURO SOCIAL",
    "SIMPLES NACIONAL",
    "TARIFA BANCARIA",
    "TARIFAS BANCARIAS",
    "TAXA ADMINISTRATIVA",
    "TAXAS BANCARIAS",
    "TREINAMENTO",
    "TRIENIO",
    "VALE TRANSPORTE",
    "VT ",
}

FORCE_INCLUDE_DRE_CODES = (
    "2.02.017.003",
    "2.05.008",
    "2.06.004",
    "2.06.014",
    "2.10",
)

FORCE_INCLUDE_BALANCETE_PATTERNS = (
    "COMISSAO PJ",
    "CURSOS E TREINAMENTOS",
    "CAJU FACILITADORES",
)


@dataclass
class DreDetail:
    code: str
    description: str
    value: float
    normalized_description: str


@dataclass
class DreCategory:
    plan01_code: str
    plan01_description: str
    plan02_code: str
    plan02_description: str
    plan03_code: str
    plan03_description: str
    total_value: float
    details: list[DreDetail] = field(default_factory=list)

    @property
    def normalized_plan03(self) -> str:
        return normalize_text(self.plan03_description)


@dataclass
class BalanceteAccount:
    clas_cta: str
    nome_cta: str
    natureza: str
    saldoatu: float
    signed_value: float

    @property
    def normalized_name(self) -> str:
        return normalize_text(self.nome_cta)


@dataclass
class BalanceteMatch:
    dre_plan03_code: str
    dre_plan03_description: str
    dre_plan02_description: str
    match_reference: str
    match_type: str
    score: float


@dataclass
class ComparisonArtifacts:
    dre_categories: list[DreCategory]
    balancete_accounts: list[BalanceteAccount]
    dre_excluded: list[dict]
    balancete_excluded: list[dict]
    period_start: str
    period_end: str
    company_id: str


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.upper().replace("&", " E ")
    text = re.sub(r"^\(?-\)\s*", "", text)
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def tokenize(value: str) -> set[str]:
    normalized_tokens = set()
    for token in normalize_text(value).split():
        if len(token) <= 1 or token in STOPWORDS:
            continue
        if token.endswith("S") and len(token) > 4:
            token = token[:-1]
        normalized_tokens.add(token)
    return normalized_tokens


def split_code_and_description(label: str | None) -> tuple[str, str]:
    if not label:
        return "", ""
    text = str(label).strip()
    match = re.match(r"^([0-9.]+)\s*-\s*(.+)$", text)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return "", text


def is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def to_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace(".", "").replace(",", ".")
    return float(text)


def should_exclude_text(*parts: str) -> bool:
    normalized_parts = [normalize_text(part) for part in parts if part]
    normalized = " ".join(normalized_parts)
    tokens = normalized.split()

    for keyword in EXCLUDED_KEYWORDS:
        keyword_tokens = normalize_text(keyword).split()
        if not keyword_tokens:
            continue

        if len(keyword_tokens) == 1:
            keyword_token = keyword_tokens[0]
            if len(keyword_token) <= 3:
                if keyword_token in tokens:
                    return True
            else:
                if any(token.startswith(keyword_token) for token in tokens):
                    return True
            continue

        if all(any(token.startswith(keyword_token) for token in tokens) for keyword_token in keyword_tokens):
            return True

    return False


def is_force_included_dre(*labels: str) -> bool:
    normalized_labels = [normalize_text(label) for label in labels if label]
    for label in normalized_labels:
        if any(label.startswith(normalize_text(code)) for code in FORCE_INCLUDE_DRE_CODES):
            return True
    return False


def is_force_included_balancete(clas_cta: str, nome_cta: str) -> bool:
    normalized_name = normalize_text(nome_cta)
    if any(normalized_name.startswith(normalize_text(pattern)) for pattern in FORCE_INCLUDE_BALANCETE_PATTERNS):
        return True
    return any(normalize_text(clas_cta).startswith(normalize_text(code)) for code in FORCE_INCLUDE_DRE_CODES)


def read_xls_bytes(source: str | bytes | BinaryIO) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, str):
        with open(source, "rb") as file:
            return file.read()
    if hasattr(source, "seek"):
        source.seek(0)
    data = source.read()
    if not isinstance(data, bytes):
        raise TypeError("Conteudo .xls invalido.")
    return data


def read_xlsx_workbook(source: str | bytes | BinaryIO):
    if isinstance(source, str):
        return load_workbook(source, data_only=True)
    if isinstance(source, bytes):
        return load_workbook(BytesIO(source), data_only=True)
    if hasattr(source, "seek"):
        source.seek(0)
    return load_workbook(source, data_only=True)


def parse_dre(source: str | bytes | BinaryIO) -> tuple[list[DreCategory], list[dict]]:
    workbook = read_xlsx_workbook(source)
    sheet = workbook[workbook.sheetnames[0]]

    current_plan01 = ""
    current_plan02 = ""
    current_plan03 = ""
    categories_by_label: dict[str, DreCategory] = {}
    pending_details: dict[str, list[DreDetail]] = {}
    excluded_rows: list[dict] = []

    for row_index in range(3, sheet.max_row + 1):
        plan01_cell = sheet.cell(row_index, 1).value
        plan02_cell = sheet.cell(row_index, 2).value
        plan03_cell = sheet.cell(row_index, 3).value
        plan04_cell = sheet.cell(row_index, 4).value
        value = sheet.cell(row_index, 5).value

        if plan01_cell and plan01_cell != "Total":
            current_plan01 = str(plan01_cell).strip()
        if plan02_cell and plan02_cell != "Total":
            current_plan02 = str(plan02_cell).strip()
        if plan03_cell and plan03_cell != "Total":
            current_plan03 = str(plan03_cell).strip()

        if not current_plan03 or not current_plan01.startswith("2"):
            continue

        if plan04_cell and str(plan04_cell).strip() == "Total" and is_number(value):
            plan01_code, plan01_description = split_code_and_description(current_plan01)
            plan02_code, plan02_description = split_code_and_description(current_plan02)
            plan03_code, plan03_description = split_code_and_description(current_plan03)

            category = DreCategory(
                plan01_code=plan01_code,
                plan01_description=plan01_description,
                plan02_code=plan02_code,
                plan02_description=plan02_description,
                plan03_code=plan03_code,
                plan03_description=plan03_description,
                total_value=to_float(value),
            )

            if (
                not is_force_included_dre(current_plan03)
                and should_exclude_text(plan02_description, plan03_description)
            ):
                excluded_rows.append(
                    {
                        "origem": "DRE",
                        "linha": row_index,
                        "conta": current_plan03,
                        "valor": category.total_value,
                        "motivo": "Filtro padrao",
                    }
                )
                continue

            category.details = pending_details.pop(current_plan03, [])
            categories_by_label[current_plan03] = category
            continue

        if not plan04_cell or not is_number(value):
            continue

        plan04_text = str(plan04_cell).strip()
        if plan04_text == "Total":
            continue

        detail_code, detail_description = split_code_and_description(plan04_text)
        if (
            not is_force_included_dre(current_plan03, plan04_text)
            and should_exclude_text(current_plan02, current_plan03, detail_description)
        ):
            excluded_rows.append(
                {
                    "origem": "DRE",
                    "linha": row_index,
                    "conta": plan04_text,
                    "valor": to_float(value),
                    "motivo": "Filtro padrao",
                }
            )
            continue

        pending_details.setdefault(current_plan03, []).append(
            DreDetail(
                code=detail_code,
                description=detail_description,
                value=to_float(value),
                normalized_description=normalize_text(detail_description),
            )
        )

    categories = list(categories_by_label.values())
    return categories, excluded_rows


def parse_balancete(source: str | bytes | BinaryIO) -> tuple[list[BalanceteAccount], list[dict], dict]:
    content = read_xls_bytes(source)
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    headers = sheet.row_values(0)

    accounts: list[BalanceteAccount] = []
    excluded_rows: list[dict] = []
    metadata = {"company_id": "", "period_start": "", "period_end": ""}

    for row_index in range(1, sheet.nrows):
        row = dict(zip(headers, sheet.row_values(row_index)))
        clas_cta = str(row.get("clas_cta", "")).strip()
        tipo_cta = str(row.get("tipo_cta", "")).strip().upper()
        nome_cta = str(row.get("nome_cta", "")).strip()
        natureza = str(row.get("natureza", "")).strip().upper()
        saldoatu = to_float(row.get("saldoatu"))

        if not metadata["company_id"]:
            metadata["company_id"] = str(row.get("cgc_emrpesa", "")).strip()
            metadata["period_start"] = str(row.get("datini", "")).strip()
            metadata["period_end"] = str(row.get("datfin", "")).strip()

        if not clas_cta.startswith("3") or tipo_cta != "A":
            continue

        if (
            not is_force_included_balancete(clas_cta, nome_cta)
            and (clas_cta.startswith(EXCLUDED_RESULT_PREFIXES) or should_exclude_text(nome_cta))
        ):
            excluded_rows.append(
                {
                    "origem": "Balancete",
                    "linha": row_index + 1,
                    "conta": nome_cta,
                    "valor": saldoatu,
                    "motivo": "Filtro padrao",
                }
            )
            continue

        signed_value = abs(saldoatu) if natureza == "D" else -abs(saldoatu)
        accounts.append(
            BalanceteAccount(
                clas_cta=clas_cta,
                nome_cta=nome_cta,
                natureza=natureza,
                saldoatu=saldoatu,
                signed_value=signed_value,
            )
        )

    return accounts, excluded_rows, metadata


def token_similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0

    left_tokens = tokenize(left)
    right_tokens = tokenize(right)
    if not left_tokens or not right_tokens:
        return 0.0

    intersection = len(left_tokens & right_tokens)
    union = len(left_tokens | right_tokens)
    jaccard = intersection / union if union else 0.0
    sequence = SequenceMatcher(None, left, right).ratio()

    score = (0.55 * sequence) + (0.45 * jaccard)

    if left_tokens <= right_tokens or right_tokens <= left_tokens:
        score += 0.12
    if left in right or right in left:
        score += 0.08

    return min(score, 1.0)


def build_exact_indexes(
    categories: list[DreCategory],
) -> tuple[dict[str, DreCategory], dict[str, list[tuple[DreCategory, DreDetail]]]]:
    category_index: dict[str, DreCategory] = {}
    detail_index: dict[str, list[tuple[DreCategory, DreDetail]]] = {}

    for category in categories:
        normalized_category = category.normalized_plan03
        if normalized_category:
            category_index.setdefault(normalized_category, category)
        for detail in category.details:
            detail_index.setdefault(detail.normalized_description, []).append((category, detail))

    return category_index, detail_index


def score_candidate(
    account: BalanceteAccount,
    category: DreCategory,
    detail: DreDetail | None = None,
) -> tuple[float, str, str]:
    target_text = detail.description if detail else category.plan03_description
    match_type = "plano_04" if detail else "plano_03"
    score = token_similarity(account.nome_cta, target_text)

    if detail and abs(abs(account.signed_value) - abs(detail.value)) <= MONEY_TOLERANCE:
        score += 0.12
    if not detail and abs(abs(account.signed_value) - abs(category.total_value)) <= MONEY_TOLERANCE:
        score += 0.15

    return min(score, 1.0), target_text, match_type


def find_best_match(
    account: BalanceteAccount,
    categories: list[DreCategory],
    category_index: dict[str, DreCategory],
    detail_index: dict[str, list[tuple[DreCategory, DreDetail]]],
) -> BalanceteMatch | None:
    normalized_name = account.normalized_name

    exact_details = detail_index.get(normalized_name, [])
    if len(exact_details) == 1:
        category, detail = exact_details[0]
        return BalanceteMatch(
            dre_plan03_code=category.plan03_code,
            dre_plan03_description=category.plan03_description,
            dre_plan02_description=category.plan02_description,
            match_reference=detail.description,
            match_type="plano_04_exato",
            score=1.0,
        )

    exact_category = category_index.get(normalized_name)
    if exact_category:
        return BalanceteMatch(
            dre_plan03_code=exact_category.plan03_code,
            dre_plan03_description=exact_category.plan03_description,
            dre_plan02_description=exact_category.plan02_description,
            match_reference=exact_category.plan03_description,
            match_type="plano_03_exato",
            score=1.0,
        )

    best_match: BalanceteMatch | None = None
    best_score = 0.0

    for category in categories:
        category_score, category_reference, category_type = score_candidate(account, category, None)
        if category_score > best_score:
            best_score = category_score
            best_match = BalanceteMatch(
                dre_plan03_code=category.plan03_code,
                dre_plan03_description=category.plan03_description,
                dre_plan02_description=category.plan02_description,
                match_reference=category_reference,
                match_type=category_type,
                score=category_score,
            )

        for detail in category.details:
            detail_score, detail_reference, detail_type = score_candidate(account, category, detail)
            if detail_score > best_score:
                best_score = detail_score
                best_match = BalanceteMatch(
                    dre_plan03_code=category.plan03_code,
                    dre_plan03_description=category.plan03_description,
                    dre_plan02_description=category.plan02_description,
                    match_reference=detail_reference,
                    match_type=detail_type,
                    score=detail_score,
                )

    if best_match and best_match.score >= 0.73:
        return best_match
    return None


def build_comparison_rows(
    dre_categories: list[DreCategory],
    balancete_accounts: list[BalanceteAccount],
) -> tuple[list[dict], list[dict], dict]:
    category_index, detail_index = build_exact_indexes(dre_categories)
    dre_rows: list[dict] = []
    balancete_rows: list[dict] = []

    category_accumulator = {
        category.plan03_code: {
            "dre_total": category.total_value,
            "plan02": category.plan02_description,
            "plan03": category.plan03_description,
            "accounts": [],
            "matched_total": 0.0,
            "match_types": set(),
        }
        for category in dre_categories
    }

    for account in balancete_accounts:
        match = find_best_match(account, dre_categories, category_index, detail_index)
        if match:
            bucket = category_accumulator[match.dre_plan03_code]
            bucket["accounts"].append(account)
            bucket["matched_total"] += account.signed_value
            bucket["match_types"].add(match.match_type)

            balancete_rows.append(
                {
                    "Clas_cta": account.clas_cta,
                    "Conta analitica": account.nome_cta,
                    "Valor balancete": account.signed_value,
                    "Natureza": account.natureza,
                    "Plano 02 DRE": match.dre_plan02_description,
                    "Plano 03 DRE": f"{match.dre_plan03_code} - {match.dre_plan03_description}".strip(" -"),
                    "Referencia encontrada": match.match_reference,
                    "Tipo de match": match.match_type,
                    "Score": round(match.score, 4),
                    "Status": "LOCALIZADO NA DRE",
                }
            )
        else:
            balancete_rows.append(
                {
                    "Clas_cta": account.clas_cta,
                    "Conta analitica": account.nome_cta,
                    "Valor balancete": account.signed_value,
                    "Natureza": account.natureza,
                    "Plano 02 DRE": "",
                    "Plano 03 DRE": "",
                    "Referencia encontrada": "",
                    "Tipo de match": "",
                    "Score": 0.0,
                    "Status": "NAO LOCALIZADO NA DRE",
                }
            )

    matched_categories = 0
    divergent_categories = 0
    missing_in_balancete = 0

    for category in dre_categories:
        bucket = category_accumulator[category.plan03_code]
        difference = bucket["matched_total"] - category.total_value
        if not bucket["accounts"]:
            status = "NAO LOCALIZADO NO BALANCETE"
            missing_in_balancete += 1
        elif abs(difference) <= MONEY_TOLERANCE:
            status = "OK"
            matched_categories += 1
        else:
            status = "DIVERGENTE"
            divergent_categories += 1

        dre_rows.append(
            {
                "Plano 02": category.plan02_description,
                "Plano 03 Codigo": category.plan03_code,
                "Plano 03 Descricao": category.plan03_description,
                "Valor DRE": category.total_value,
                "Valor Balancete encontrado": bucket["matched_total"],
                "Diferenca": difference,
                "Status": status,
                "Contas do balancete": " | ".join(account.nome_cta for account in bucket["accounts"]),
                "Tipos de match": ", ".join(sorted(bucket["match_types"])),
                "Detalhes DRE": " | ".join(detail.description for detail in category.details),
            }
        )

    missing_in_dre = sum(1 for row in balancete_rows if row["Status"] == "NAO LOCALIZADO NA DRE")

    summary = {
        "categorias_dre_analisadas": len(dre_categories),
        "contas_balancete_analisadas": len(balancete_accounts),
        "categorias_ok": matched_categories,
        "categorias_divergentes": divergent_categories,
        "categorias_nao_localizadas_no_balancete": missing_in_balancete,
        "contas_nao_localizadas_na_dre": missing_in_dre,
    }
    return dre_rows, balancete_rows, summary


def compare_reports(
    balancete_source: str | bytes | BinaryIO,
    dre_source: str | bytes | BinaryIO,
) -> tuple[list[dict], list[dict], dict, ComparisonArtifacts]:
    dre_categories, dre_excluded = parse_dre(dre_source)
    balancete_accounts, balancete_excluded, metadata = parse_balancete(balancete_source)
    dre_rows, balancete_rows, summary = build_comparison_rows(dre_categories, balancete_accounts)

    summary["registros_dre_filtrados"] = len(dre_excluded)
    summary["registros_balancete_filtrados"] = len(balancete_excluded)

    artifacts = ComparisonArtifacts(
        dre_categories=dre_categories,
        balancete_accounts=balancete_accounts,
        dre_excluded=dre_excluded,
        balancete_excluded=balancete_excluded,
        period_start=metadata["period_start"],
        period_end=metadata["period_end"],
        company_id=metadata["company_id"],
    )
    return dre_rows, balancete_rows, summary, artifacts


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def get_status_fill(status: str | None):
    mapping = {
        "OK": FILL_OK,
        "DIVERGENTE": FILL_DIVERGENT,
        "NAO LOCALIZADO NO BALANCETE": FILL_MISSING,
        "LOCALIZADO NA DRE": FILL_FOUND,
        "NAO LOCALIZADO NA DRE": FILL_NOT_FOUND_DRE,
    }
    return mapping.get(str(status or "").strip(), FILL_NEUTRAL)


def style_title_cell(cell) -> None:
    cell.fill = FILL_TITLE
    cell.font = FONT_WHITE_BOLD
    cell.alignment = WRAP_TOP


def style_section_cell(cell) -> None:
    cell.fill = FILL_SECTION
    cell.font = FONT_BOLD
    cell.alignment = WRAP_TOP


def style_header_cell(cell) -> None:
    cell.fill = FILL_HEADER
    cell.font = FONT_BOLD
    cell.alignment = WRAP_TOP


def style_data_row(sheet, row_index: int, total_columns: int, fill) -> None:
    for column_index in range(1, total_columns + 1):
        cell = sheet.cell(row_index, column_index)
        cell.fill = fill
        cell.alignment = WRAP_TOP
        if isinstance(cell.value, float):
            cell.number_format = '#,##0.00'


def write_table(sheet, rows: list[dict]) -> None:
    write_table_at(sheet, rows, start_row=1)


def write_table_at(sheet, rows: list[dict], start_row: int = 1) -> None:
    if not rows:
        sheet.cell(start_row, 1).value = "Sem registros"
        sheet.cell(start_row, 1).alignment = WRAP_TOP
        autosize_columns(sheet)
        return

    headers = list(rows[0].keys())
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(start_row, column_index).value = header

    for cell in sheet[start_row]:
        style_header_cell(cell)

    current_row = start_row + 1
    for row in rows:
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(current_row, column_index)
            cell.value = row.get(header, "")
            cell.alignment = WRAP_TOP
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'
        current_row += 1

    autosize_columns(sheet)


def build_report_workbook(
    dre_rows: list[dict],
    balancete_rows: list[dict],
    summary: dict,
    artifacts: ComparisonArtifacts,
    balancete_filename: str,
    dre_filename: str,
) -> bytes:
    workbook = Workbook()

    comparative_sheet = workbook.active
    comparative_sheet.title = "Comparativo"

    comparative_rows = []
    for row in dre_rows:
        comparative_rows.append(
            {
                "Origem": "DRE",
                "Grupo": row["Plano 02"],
                "Codigo": row["Plano 03 Codigo"],
                "Descricao": row["Plano 03 Descricao"],
                "Valor origem": row["Valor DRE"],
                "Valor encontrado no outro relatorio": row["Valor Balancete encontrado"],
                "Diferenca": row["Diferenca"],
                "Status": row["Status"],
                "Referencia cruzada": row["Contas do balancete"],
                "Observacoes": row["Detalhes DRE"],
            }
        )

    comparative_sheet["A1"] = "Legenda dos status"
    style_title_cell(comparative_sheet["A1"])
    comparative_sheet["A2"] = "OK"
    comparative_sheet["B2"] = "Categoria da DRE localizada no balancete e com valor total compativel."
    comparative_sheet["A3"] = "DIVERGENTE"
    comparative_sheet["B3"] = "Categoria da DRE localizada no balancete, mas com diferenca de valor."
    comparative_sheet["A4"] = "NAO LOCALIZADO NO BALANCETE"
    comparative_sheet["B4"] = "Categoria da DRE sem conta analitica equivalente localizada no balancete."
    comparative_sheet["A6"] = "Tabela comparativa"
    style_title_cell(comparative_sheet["A6"])
    for row_index in range(2, 5):
        comparative_sheet[f"A{row_index}"].fill = get_status_fill(comparative_sheet[f"A{row_index}"].value)
        comparative_sheet[f"A{row_index}"].font = FONT_BOLD
        comparative_sheet[f"A{row_index}"].alignment = WRAP_TOP
        comparative_sheet[f"B{row_index}"].fill = get_status_fill(comparative_sheet[f"A{row_index}"].value)
        comparative_sheet[f"B{row_index}"].alignment = WRAP_TOP
    write_table_at(comparative_sheet, comparative_rows, start_row=10)
    comparative_sheet.freeze_panes = "A11"

    if comparative_rows:
        comparative_headers = list(comparative_rows[0].keys())
        comparative_status_column_index = comparative_headers.index("Status") + 1
        for row_index in range(11, comparative_sheet.max_row + 1):
            status_value = comparative_sheet.cell(row_index, comparative_status_column_index).value
            style_data_row(
                comparative_sheet,
                row_index=row_index,
                total_columns=len(comparative_headers),
                fill=get_status_fill(status_value),
            )

    summary_sheet = workbook.create_sheet("Resumo")
    summary_sheet["A1"] = "Status geral da comparacao"
    style_title_cell(summary_sheet["A1"])

    summary_rows = [
        ("Balancete", balancete_filename),
        ("DRE", dre_filename),
        ("CNPJ", artifacts.company_id),
        ("Periodo inicial", artifacts.period_start),
        ("Periodo final", artifacts.period_end),
        ("Categorias DRE analisadas", summary["categorias_dre_analisadas"]),
        ("Contas balancete analisadas", summary["contas_balancete_analisadas"]),
        ("Categorias OK", summary["categorias_ok"]),
        ("Categorias divergentes", summary["categorias_divergentes"]),
        ("Categorias nao localizadas no balancete", summary["categorias_nao_localizadas_no_balancete"]),
        ("Contas nao localizadas na DRE", summary["contas_nao_localizadas_na_dre"]),
        ("Registros DRE filtrados", summary["registros_dre_filtrados"]),
        ("Registros balancete filtrados", summary["registros_balancete_filtrados"]),
    ]

    for index, (label, value) in enumerate(summary_rows, start=3):
        summary_sheet[f"A{index}"] = label
        summary_sheet[f"B{index}"] = value
        style_section_cell(summary_sheet[f"A{index}"])
        summary_sheet[f"B{index}"].alignment = WRAP_TOP
        if isinstance(value, (int, float)):
            summary_sheet[f"B{index}"].number_format = '#,##0.00'

    summary_sheet["A17"] = "Observacao"
    style_section_cell(summary_sheet["A17"])
    summary_sheet["B17"] = (
        "O comparativo usa o Plano de conta nº. 03 da DRE e tenta localizar contas analiticas "
        "do balancete por nome, primeiro no Plano 04 e depois no Plano 03."
    )
    summary_sheet["B18"] = (
        "Filtros padrao removem receitas, impostos, contas de folha, financeiras e contas sem nota, "
        "com excecoes explicitas para categorias liberadas nesta analise."
    )
    summary_sheet["B17"].alignment = WRAP_TOP
    summary_sheet["B18"].alignment = WRAP_TOP
    detail_sections = [
        (
            "Categorias da DRE analisadas",
            [f"{item.plan03_code} - {item.plan03_description}" for item in artifacts.dre_categories],
        ),
        (
            "Contas do balancete analisadas",
            [f"{item.clas_cta} - {item.nome_cta}" for item in artifacts.balancete_accounts],
        ),
        (
            "Registros DRE filtrados e excluidos",
            [
                f"Linha {item['linha']}: {item['conta']} | valor={item['valor']} | motivo={item['motivo']}"
                for item in artifacts.dre_excluded
            ],
        ),
        (
            "Registros balancete filtrados e excluidos",
            [
                f"Linha {item['linha']}: {item['conta']} | valor={item['valor']} | motivo={item['motivo']}"
                for item in artifacts.balancete_excluded
            ],
        ),
        (
            "Categorias OK",
            [
                f"{item['Plano 03 Codigo']} - {item['Plano 03 Descricao']}"
                for item in dre_rows
                if item["Status"] == "OK"
            ],
        ),
        (
            "Categorias divergentes",
            [
                f"{item['Plano 03 Codigo']} - {item['Plano 03 Descricao']} | DRE={item['Valor DRE']} | Balancete={item['Valor Balancete encontrado']} | Dif={item['Diferenca']}"
                for item in dre_rows
                if item["Status"] == "DIVERGENTE"
            ],
        ),
        (
            "Categorias da DRE nao localizadas no balancete",
            [
                f"{item['Plano 03 Codigo']} - {item['Plano 03 Descricao']}"
                for item in dre_rows
                if item["Status"] == "NAO LOCALIZADO NO BALANCETE"
            ],
        ),
        (
            "Contas do balancete localizadas na DRE",
            [
                f"{item['Clas_cta']} - {item['Conta analitica']} | referencia={item['Plano 03 DRE']}"
                for item in balancete_rows
                if item["Status"] == "LOCALIZADO NA DRE"
            ],
        ),
        (
            "Contas do balancete nao localizadas na DRE",
            [
                f"{item['Clas_cta']} - {item['Conta analitica']} | valor={item['Valor balancete']}"
                for item in balancete_rows
                if item["Status"] == "NAO LOCALIZADO NA DRE"
            ],
        ),
    ]

    start_row = 20
    for title, values in detail_sections:
        summary_sheet[f"A{start_row}"] = title
        style_section_cell(summary_sheet[f"A{start_row}"])
        if values:
            for offset, value in enumerate(values, start=1):
                summary_sheet[f"A{start_row + offset}"] = value
                summary_sheet[f"A{start_row + offset}"].alignment = WRAP_TOP
            start_row += len(values) + 2
        else:
            summary_sheet[f"A{start_row + 1}"] = "Sem registros"
            summary_sheet[f"A{start_row + 1}"].alignment = WRAP_TOP
            start_row += 3

    autosize_columns(summary_sheet)

    balancete_detail_rows = []
    for row in balancete_rows:
        balancete_detail_rows.append(
            {
                "Origem": "Balancete",
                "Grupo": row["Plano 02 DRE"],
                "Codigo": row["Clas_cta"],
                "Descricao": row["Conta analitica"],
                "Valor origem": row["Valor balancete"],
                "Valor encontrado no outro relatorio": row["Valor balancete"] if row["Status"] == "LOCALIZADO NA DRE" else "",
                "Diferenca": "",
                "Status": row["Status"],
                "Referencia cruzada": row["Plano 03 DRE"],
                "Observacoes": row["Referencia encontrada"],
            }
        )

    balancete_sheet = workbook.create_sheet("Contas Balancete")
    balancete_sheet["A1"] = "Legenda dos status"
    style_title_cell(balancete_sheet["A1"])
    balancete_sheet["A2"] = "LOCALIZADO NA DRE"
    balancete_sheet["B2"] = "Conta analitica do balancete localizada em categoria equivalente da DRE."
    balancete_sheet["A3"] = "NAO LOCALIZADO NA DRE"
    balancete_sheet["B3"] = "Conta analitica do balancete sem categoria equivalente localizada na DRE."
    balancete_sheet["A5"] = "Tabela das contas do balancete"
    style_title_cell(balancete_sheet["A5"])
    for row_index in range(2, 4):
        balancete_sheet[f"A{row_index}"].fill = get_status_fill(balancete_sheet[f"A{row_index}"].value)
        balancete_sheet[f"A{row_index}"].font = FONT_BOLD
        balancete_sheet[f"A{row_index}"].alignment = WRAP_TOP
        balancete_sheet[f"B{row_index}"].fill = get_status_fill(balancete_sheet[f"A{row_index}"].value)
        balancete_sheet[f"B{row_index}"].alignment = WRAP_TOP
    write_table_at(balancete_sheet, balancete_detail_rows, start_row=10)
    balancete_sheet.freeze_panes = "A11"

    if balancete_detail_rows:
        balancete_headers = list(balancete_detail_rows[0].keys())
        balancete_status_column_index = balancete_headers.index("Status") + 1
        for row_index in range(11, balancete_sheet.max_row + 1):
            status_value = balancete_sheet.cell(row_index, balancete_status_column_index).value
            style_data_row(
                balancete_sheet,
                row_index=row_index,
                total_columns=len(balancete_headers),
                fill=get_status_fill(status_value),
            )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
